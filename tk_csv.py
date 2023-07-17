#!/usr/bin/python
# -*- coding: UTF-8 -*-

import os
import csv
import openpyxl
import pandas as pd
import shutil
import tkinter as tk
from tkinterdnd2 import DND_FILES, TkinterDnD
from PIL import Image, ImageTk 
from tkinter import messagebox
from datetime import date
from openpyxl import load_workbook

import win32com.client as win32
import win32com
# print(win32com.__gen_path__)

class report_pheniX():

    def __init__(self):

        self.file_found=False
        self.result = ""

        self.root = TkinterDnD.Tk()
        # screen_width = self.root.winfo_screenwidth()
        # screen_height = self.root.winfo_screenheight()
        screen_width = 1400
        screen_height = 850
        window_width = int(screen_width / 1.5)
        window_height = int(screen_height / 1.2)
        self.root.geometry(f"{window_width}x{window_height}+{int((screen_width - window_width) / 1.2)}+{int((screen_height - window_height) / 1.2)}")

        self.root.configure(bg='#FFFDD0')

        image = Image.open("SINBON LOGO-01.jpg")

        # 縮小圖片尺寸為原來的一半
        resized_image = image.resize((image.size[0]//3, image.size[1]//3))

        # 將圖片轉換為TK格式
        tk_image = ImageTk.PhotoImage(resized_image)

        # 创建标签并设置位置
        label = tk.Label(self.root, image=tk_image)
        label.place(x=410, y=20)

        checkbox_frame = tk.Frame(self.root, bg='#FFFFFF',relief=tk.GROOVE,borderwidth=2,width=450, height=200)
        checkbox_frame.place(x=400, y=170)


        
        self.filename=""
        self.root.drop_target_register(DND_FILES)
        self.root.dnd_bind('<<Drop>>',self.on_drop)
        self.label_filename =tk.Label(self.root,text="拖曳CSV檔到此視窗",font=('Helvetica', 18), bg='#FFFFFF')
        self.label_filename.place(x=430,y=250)
        self.test_time_csv =tk.Label(self.root,text="",font=('Helvetica', 18), bg='#FFFFFF')
        self.test_time_csv.place(x=430,y=300)

        self.max_voltage_label = tk.Label(self.root,relief="groove",bg='#C0C0C0')
        self.max_voltage_label.place(x=400,y=380,width=280, height=100)

        self.max_voltage = tk.Label(self.root,text="最大電壓=",font=('Helvetica', 20), bg='#C0C0C0')
        self.max_voltage.place(x=430,y=410)

        start_button = tk.Button(self.root,command=self.start_button,text="Start",font=('Helvetica', 18),bg="#87CEFA")
        start_button.place(x=30,y=620,width=160,height=70)


        self.entry_lable = tk.Label(self.root,relief="groove",bg='#B0C4DE')
        self.entry_lable.place(x=20,y=150,width=320, height=450)
        employee_number_frame = tk.Label(self.root,relief="groove", bg='#87CEFA')
        employee_number_frame.place(x=20,y=30,width=320, height=100)

        self.employee_number_label =tk.Label(self.root,text="員工工號",font=('Helvetica', 20), bg='#87CEFA')
        self.employee_number_label.place(x=120,y=40)

        self.employee_number_entry = tk.Entry(self.root,font=('Helvetica', 18))
        self.employee_number_entry.place(x=40,y=80, width=280, height=35)

        self.label_KVset_number=tk.Label(self.root,text="設定電壓(KV)",font=('Helvetica', 20), bg='#B0C4DE')
        self.label_KVset_number.place(x=110,y=170)  

        self.input_entry_4 = tk.Entry(self.root,font=('Helvetica', 18))
        self.input_entry_4.place(x=40,y=220, width=280, height=35)
                
        self.label_work_order=tk.Label(self.root,text="工單號碼",font=('Helvetica', 20), bg='#B0C4DE')
        self.label_work_order.place(x=120,y=270)

        self.input_entry = tk.Entry(self.root,font=('Helvetica', 18))
        self.input_entry.place(x=40,y=310, width=280, height=35)

        self.label_customer_part_number=tk.Label(self.root,text="BID",font=('Helvetica', 20), bg='#B0C4DE')
        self.label_customer_part_number.place(x=150,y=355)        
  
        self.input_entry_2 = tk.Entry(self.root,font=('Helvetica', 18))
        self.input_entry_2.place(x=40,y=395, width=280, height=35)

        self.label_sinbon_part_number=tk.Label(self.root,text="信邦料號",font=('Helvetica', 20), bg='#B0C4DE')
        self.label_sinbon_part_number.place(x=120,y=440)  

        self.input_entry_3 = tk.Entry(self.root,font=('Helvetica', 18))
        self.input_entry_3.place(x=40,y=480, width=280, height=35)  

        


        self.result_show = tk.Label(self.root,text="wait",font=('Helvetica', 35), bg='#F0F8FF',relief=tk.GROOVE)
        self.result_show.place(x=420,y=520,width=400, height=80)

        def on_employee_number(event):
            # 在 Entry 1 中按下 Enter 键时触发的事件处理函数
            # 将焦点移动到 Entry 2
            self.input_entry_4.focus_set()
        def on_input_entry_4(event):
            # 在 Entry 1 中按下 Enter 键时触发的事件处理函数
            # 将焦点移动到 Entry 2
            self.input_entry.focus_set()           
        def on_enter1(event):
            # 在 Entry 1 中按下 Enter 键时触发的事件处理函数
            # 将焦点移动到 Entry 2
            self.input_entry_2.focus_set()


        def on_enter2(event):
            # 在 Entry 2 中按下 Enter 键时触发的事件处理函数
            # 将焦点移动到 Entry 3
            self.input_entry_3.focus_set()

        def on_enter3(event):
            # 在 Entry 3 中按下 Enter 键时触发的事件处理函数
            # 将焦点移动到 Entry 1
            self.input_entry.focus_set()
        self.employee_number_entry.bind('<Return>',on_employee_number)
        self.input_entry_4.bind('<Return>',on_input_entry_4)   
        self.input_entry.bind('<Return>', on_enter1)
        self.input_entry_2.bind('<Return>', on_enter2)
        self.input_entry_3.bind('<Return>', on_enter3)



        # label_a = tk.Label(self.root, text='',bg='#ADD8E6',font=('Arial',15,"bold"),relief="sunken")
        # label_a.place(x=30,y=120, width=150, height=30)

        # label_b = tk.Label(self.root, text='',bg='#ADD8E6',font=('Arial',15,"bold"),relief="sunken")
        # label_b.place(x=30,y=160, width=150, height=30)

        self.root.title("絕緣測試")
        self.root.mainloop()



    def on_drop(self,event):
        self.filepath =event.data
        self.filename =self.filepath.split('/')[-1]
        self.filename = self.filename.replace('.csv', '')
        self.label_filename.config(text="CSV檔名:"+self.filename)

        print(self.filepath,self.filename)
        test_time=self.read_csv_cell(self.filepath)
        self.test_time_csv.config(text="測試時間:"+test_time)
        self.file_found=True

    def get_entry(self):

        self.input_entry_get = str(self.input_entry.get())
        self.input_entry_get_2 = self.input_entry_2.get()
        self.input_entry_get_2 = self.input_entry_get_2.strip()
        self.input_entry_get_3 = self.input_entry_3.get()
        self.input_entry_get_3 = self.input_entry_get_3.strip()
        self.input_entry_get_4 = self.input_entry_4.get()
        self.input_entry_get_4 = self.input_entry_get_4.strip()
        if self.input_entry_get_4 != "":
            self.input_entry_get_4 = int(self.input_entry_get_4)
        self.employee_number_entry_get = self.employee_number_entry.get()
        self.employee_number_entry_get= self.employee_number_entry_get.strip()
        if self.employee_number_entry_get != "":
            self.employee_number_entry_get = int(self.employee_number_entry_get)
        len_input_entry_get_2 = len(self.input_entry_get_2)
        print(len_input_entry_get_2)
        self.entry_status = False
        if len_input_entry_get_2 < 16  :
            messagebox.showinfo("ERROR","請檢查客戶料號字數")
            self.result_show.config(text="FAIL",bg="#FF0000")
        elif len_input_entry_get_2 > 18 :
            messagebox.showinfo("ERROR","請檢查客戶料號字數")
            self.result_show.config(text="FAIL",bg="#FF0000")
        else:
            
            if self.input_entry_get =="":
                messagebox.showinfo("ERROR","工單號碼不得為空")
                self.result_show.config(text="FAIL",bg="#FF0000")
            elif self.input_entry_get_2 =="":
                messagebox.showinfo("ERROR","客戶料號不得為空")
                self.result_show.config(text="FAIL",bg="#FF0000")
            elif self.input_entry_get_3 == "":
                messagebox.showinfo("ERROR","信邦料號不得為空")
                self.result_show.config(text="FAIL",bg="#FF0000")    
            elif self.input_entry_get_4 =="":
                messagebox.showinfo("ERROR","電壓設定不得為空")
                self.result_show.config(text="FAIL",bg="#FF0000")
            elif  self.employee_number_entry_get == "":
                messagebox.showinfo("ERROR","員工工號不得為空")
                self.result_show.config(text="FAIL",bg="#FF0000")
            elif self.filename == "":
                messagebox.showinfo("ERROR","請拖曳需轉換報表")
                self.result_show.config(text="FAIL",bg="#FF0000")
            else :
                self.entry_status = True

    def start_button(self):
        self.result_show.config(text="Running",bg="#FFFF00")
        self.max_voltage.config(text="最大電壓=")
        # try:
        self.get_entry()
        if self.entry_status :
            self.find_csv_file()
        # except:
        #     messagebox.showinfo("ERROR","請檢查檔案與輸入資料")


    def read_csv_cell(self,csv_path):
        with open(csv_path, 'r') as file:
            reader = csv.reader(file)
            # 跳過前三行
            for _ in range(3):
                next(reader)
            # 讀取A4單元格值
            row = next(reader)
            cell_value = row[0]
            prefix = "Start time: "
            if cell_value.startswith(prefix):
                return cell_value[len(prefix):]
            # return cell_value

    def find_csv_file(self):
        
        # 設定當前工作目錄為程式檔案所在目錄
        current_dir = os.getcwd()
        os.chdir(current_dir)

        files = os.listdir('./file_a')
        num_files = len(files)
        current_date = date.today()
        formatted_date = current_date.strftime("%Y%m%d") 

        # 如果找到.csv檔案，則讀取檔案並且逐行紀錄於list內，並且將數值紀錄於Excel檔案中
        if self.file_found:
            print(self.filepath)
            data = self.read_csv_file(self.filepath)

            #這裡的 os.path.join() 方法是用來建立跨平台的路徑，可以根據當前作業系統自動調整路徑分隔符號。例如，在 Windows 系統中路徑分隔符號為反斜線 \，而在 Linux 或 macOS 系統中為正斜線 /。
            result_dir = os.path.join(current_dir, 'Result')
            os.makedirs(result_dir, exist_ok=True)
            folder_path =os.path.join(result_dir, formatted_date)

            result_file = os.path.join(folder_path, self.filename+'.xlsx')
                    # 检查文件夹是否存在
            if not os.path.exists(folder_path):
                # 创建文件夹
                os.makedirs(folder_path)
                # print("文件夹已创建：" + folder_path)
            else:
                # print("文件夹已存在：" + folder_path)
                pass
            self.template = os.path.join(current_dir, 'High_voltage_test_report_sample.xlsx')
            template = openpyxl.load_workbook(self.template)

            # 複製模板檔案
            shutil.copyfile(self.template, result_file)

            # 開啟複製後的 Excel 檔案
            workbook = openpyxl.load_workbook(result_file)
            # worksheet = workbook.active
            worksheet = workbook.create_sheet('Sheet2')
            worksheet = workbook.worksheets[2]
            # 將資料寫入指定的工作表中
            for row in data:
                worksheet.append(row)

            # 儲存並關閉Excel寫入器
            workbook.save(result_file)
            workbook.close()

            self.result_file_path =result_file
            # 打开Excel文件
            self.workbook = openpyxl.load_workbook(self.result_file_path)

            # 选择第一个工作表
            self.worksheet = self.workbook.worksheets[2]
            # print(self.worksheet)
            self.process_excel_file(self.result_file_path)  #處理報表內數值格式
            self.process_excel_file2()
            self.Calculate_voltage_value()

            
            print('檔案已儲存於' + self.result_file_path)

            print(self.result)
            self.excel_to_pdf(self.result_file_path, folder_path+"\\"+self.filename+".pdf",folder_path)
            self.clear_entries()

        else:
            print('找不到CSV檔案')
            self.worksheet['K2'].value = "FAIL"
            self.result_show.config(text="FAIL",bg="#FF0000")
        original_dir = os.path.join(current_dir, 'Original')
        os.makedirs(result_dir, exist_ok=True)
        # print(original_dir)
        # filename = os.path.join('./file_a', file)

        shutil.move(self.filepath, original_dir)


    # 定義函式以讀取CSV檔案
    def read_csv_file(self,filename):
        data = []
        with open(filename, newline='') as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                row_float = []
                for val in row:
                    if "Inf" not in val:

                        try:
                            val_float = float(val)  # 轉換成 float
                            if val_float.is_integer():  # 如果是整數就轉換成整數
                                val_float = int(val_float)
                            row_float.append(val_float)
                        except ValueError:  # 如果無法轉換成 float，就保持原樣
                            row_float.append(val)
                    else:
                        val = "Infinity"
                        row_float.append(val)
                data.append(row_float)
        return data


    def Calculate_voltage_value(self):
        self.worksheet = self.workbook.worksheets[2]

        target_value = self.input_entry_get_4
        cell_addresses = []
        find = False
        for row in self.worksheet.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value > target_value:
                    
                    cell_addresses.append(cell.coordinate)
                    find = True
                    break
            if find == True :
                break   
        if cell_addresses:
            print("找到值為", target_value, "的單元格，位於", cell_addresses)
            first_cell_address = cell_addresses[0]
            value = self.worksheet[first_cell_address].value
            converted_cell = first_cell_address.replace("B", "A")
            # print(self.process_cells(converted_cell))  # 输出：A10
            self.time_data_list = self.process_cells(converted_cell) 
            self.calculate_average(self.time_data_list)
            
            
        else:
            self.result = False
            print("未找到值為", target_value, "的單元格")
            self.worksheet = self.workbook.worksheets[1]
            
            
            self.worksheet['F2'].value = ""
        self.order_number_part_number()  #將工號等資料填入表格
        self.worksheet['L2'].value = self.employee_number_entry_get
        self.worksheet = self.workbook.worksheets[1]
        if self.result == True :
            self.worksheet['K2'].value = "PASS"
            self.result_show.config(text="PASS",bg="#00FF00")
        elif self.result == False :
            self.result_show.config(text="FAIL",bg="#FF0000")
            self.worksheet['K2'].value = "FAIL"
        else:
            print("no result")
        self.workbook.save(self.result_file_path)

        # self.worksheet = self.workbook.worksheets[1]
        # # 将最大值写入D5单元格
        # self.worksheet['G2'].value = max_value_d
        # # print(max_value_d)
        # self.worksheet = self.workbook.worksheets[2]
        # self.process_excel_file2()

    def process_cells(self,cell_address):
        column_letter = cell_address[0]  # 提取單元格地址中的列字母（例如：B10 中的 B）
        start_row = int(cell_address[1:])  # 提取單元格地址中的起始行號（例如：B10 中的 10）
        target_value = int(self.worksheet[cell_address].value) + 60  # 讀取單元格的值並加上60

        cell_addresses = []
        for row in self.worksheet.iter_rows(min_row=start_row, min_col=1, max_col=1):
            
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value < target_value:
                    cell_addresses.append(cell.coordinate)
                    

        return cell_addresses
    def calculate_average(self,time_data_list):
        total = 0
  
        len_data = len(time_data_list)


        for i in range(len_data):
            time_cell_address = self.time_data_list[i]
            converted_cell = time_cell_address.replace("A", "E")
            
            value = self.worksheet[converted_cell].value
            total += value 
            # print(value)

        average_resistance = total / len_data
        print(average_resistance)
        self.worksheet = self.workbook.worksheets[1]
        # 将最大值写入D5单元格
        self.worksheet['F2'].value = average_resistance


    def order_number_part_number(self):
        self.worksheet = self.workbook.worksheets[1]
        try:
            self.worksheet['B2'].value = self.input_entry_get
            self.worksheet['C2'].value = self.input_entry_get_2
            self.worksheet['D2'].value = self.input_entry_get_3
        except:
            messagebox.showinfo("ERROR","請檢查輸入資料")
    def clear_entries(self):
        self.input_entry.delete(0, tk.END)
        self.input_entry_2.delete(0, tk.END)
        self.input_entry_3.delete(0, tk.END)

    def process_excel_file(self, filepath):
        # 处理D列及其之后的单元格
        row_index = 7
        column_index = 2
        while True:
            cell_value = self.worksheet.cell(row=row_index, column=column_index).value
  
            if cell_value is None:
                break

            # 如果单元格所在列是D列并且单元格的值是数字类型，就将其乘以1e6，否则跳过
            if isinstance(cell_value, (int, float)) and column_index == 2:
                processed_value = cell_value /1000
                self.worksheet.cell(row=row_index, column=column_index).value = processed_value

            # 移动到下一个单元格
            column_index += 1
            if self.worksheet.cell(row=row_index, column=column_index).value is None:
                row_index += 1
                column_index = 2


        # 找到D列及其之后的单元格中的最大值
        max_value_d = 0
        for cell in self.worksheet['B7':'B' + str(row_index - 1)]:
            for value in cell:
                if isinstance(value.value, (int, float)) and value.value > max_value_d:
                    max_value_d = value.value
        
        self.worksheet = self.workbook.worksheets[1]
        # 将最大值写入D5单元格
        self.worksheet['G2'].value = max_value_d
        max_voltage = round(max_value_d, 2)
        self.max_voltage.config(text="最大電壓="+str(max_voltage))
        # print(max_value_d)
        self.worksheet = self.workbook.worksheets[2]
        

    def process_excel_file2(self):
        # 处理E列及其之后的单元格
        row_index_e = 7
        column_index_e = 5
        while True:
            cell_value_E = self.worksheet.cell(row=row_index_e, column=column_index_e).value
            if cell_value_E is None:
                break
             
            # 如果单元格的值是数字类型，就将其除以1e9，否则跳过
            if isinstance(cell_value_E, (int, float)) and column_index_e == 5 :
                processed_value1 = cell_value_E / 1000000000
            
                self.worksheet.cell(row=row_index_e, column=column_index_e).value = processed_value1
            # print(processed_value)
            # 移动到下一个单元格
            column_index_e += 1
            if self.worksheet.cell(row=row_index_e, column=column_index_e).value is None:
                row_index_e += 1
                column_index_e = 5

        # 找到E列及其之后的单元格中的最大值
        max_value_e = 0
        for cell in self.worksheet['E7':'E' + str(row_index_e - 1)]:
            for value in cell:
                if isinstance(value.value, (int, float)) and value.value > max_value_e:
                    max_value_e = value.value

        self.worksheet = self.workbook.worksheets[1]
        if max_value_e > 100:
            
            self.result = True
        else:
            
            self.result = False
        # 将最大值写入F2单元格
        self.worksheet['F2'].value = max_value_e
    def excel_to_pdf(self,excel_file, pdf_file,save_path):
        # excel_app = win32.gencache.EnsureDispatch('Excel.Application')
        excel_app = win32com.client.DispatchEx('Excel.Application')
        # 打开Excel文件
        workbook = excel_app.Workbooks.Open(excel_file)

        # 构建完整的文件路径
        pdf_path = os.path.join(save_path, pdf_file)

        # 保存为PDF
        workbook.ExportAsFixedFormat(0, pdf_path)

        # 关闭Excel文件
        workbook.Close()

        # 退出Excel应用程序
        excel_app.Quit()

        print("PDF文件已保存：" + pdf_path)
    


if __name__ =="__main__":
    
    r = report_pheniX()






